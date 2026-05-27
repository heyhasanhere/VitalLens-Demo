"""
Per-timestamp video analysis — streams video frame-by-frame (constant ~8 MB RAM).

Usage:
  python analyze_recording.py --video path.mp4 --gt "18:94,48:80,1m18s:79,..."
"""
from __future__ import annotations

import argparse
import sys
from collections import deque
from pathlib import Path

import cv2
import numpy as np
import onnxruntime as ort
from scipy.signal import butter, filtfilt, welch

_REPO_ROOT = Path(__file__).resolve().parent.parent
ONNX_PATH  = _REPO_ROOT / "weights" / "vitallens_rppg.onnx"
CLIP_LEN   = 160
STRIDE     = 80
INPUT_SIZE = 128
TARGET_FPS = 30.0


# ---------------------------------------------------------------------------
# Signal helpers
# ---------------------------------------------------------------------------

def bvp_to_hr(bvp: np.ndarray, fps: float = TARGET_FPS) -> float:
    if len(bvp) < 10:
        return 0.0
    nyq = fps / 2.0
    lo, hi = max(0.67 / nyq, 1e-4), min(3.0 / nyq, 0.9999)
    if lo >= hi:
        return 0.0
    b, a = butter(3, [lo, hi], btype="band")
    sig  = filtfilt(b, a, bvp)
    # nfft=2048 zero-pads for sub-bin interpolation (true resolution still 1/clip_len)
    freqs, psd = welch(sig, fs=fps, nperseg=min(len(sig), 160), nfft=2048)
    band = (freqs >= 0.67) & (freqs <= 3.0)
    if not band.any():
        return 0.0
    peak_idx = int(np.where(band)[0][psd[band].argmax()])
    f_peak   = freqs[peak_idx]

    # Harmonic suppression: real PPG has strong 2nd harmonic; if dominant peak
    # looks like a harmonic (> 1.3 Hz ≈ 78 BPM), check whether f/2 has
    # meaningful power (≥35% of the dominant peak). If so, prefer the fundamental.
    if f_peak > 1.3:
        f_half    = f_peak / 2.0
        half_idx  = int(np.argmin(np.abs(freqs - f_half)))
        if freqs[half_idx] >= 0.67 and psd[half_idx] >= 0.35 * psd[peak_idx]:
            sub_band = (freqs >= 0.67) & (freqs <= f_peak * 0.6)
            if sub_band.any():
                f_peak = freqs[sub_band][psd[sub_band].argmax()]

    return float(f_peak * 60.0)


def _bvp_snr(bvp: np.ndarray, fps: float = TARGET_FPS) -> float:
    """Peak / median power ratio in cardiac band — same metric as the live backend."""
    if len(bvp) < 10:
        return 0.0
    freqs, psd = welch(bvp, fs=fps, nperseg=min(len(bvp), 160), nfft=2048)
    band = (freqs >= 0.67) & (freqs <= 3.0)
    if not band.any():
        return 0.0
    peak   = psd[band].max()
    median = np.median(psd[band])
    return float(peak / (median + 1e-9))


# ---------------------------------------------------------------------------
# Streaming analysis  (constant memory: 161 × 128×128×3 ≈ 8 MB in buffer)
# ---------------------------------------------------------------------------

def analyze(video_path: str, ground_truth: list[tuple[float, float]] | None, onnx_path: Path):
    sess    = ort.InferenceSession(str(onnx_path), providers=["CPUExecutionProvider"])
    in_name = sess.get_inputs()[0].name

    # Haar face detector
    haar = cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
    det  = cv2.CascadeClassifier(haar)
    if det.empty():
        det = None
        print("Warning: Haar cascade not found — processing full frame")

    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        sys.exit(f"Cannot open: {video_path}")

    raw_fps  = cap.get(cv2.CAP_PROP_FPS) or TARGET_FPS
    n_raw    = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    # Subsample so effective fps ≈ 30
    fps_stride = max(1, round(raw_fps / TARGET_FPS))
    eff_fps    = raw_fps / fps_stride
    duration   = n_raw / raw_fps
    n_eff_est  = n_raw // fps_stride

    print(f"Video      : {video_path}")
    print(f"FPS        : {raw_fps:.1f} → {eff_fps:.1f} Hz effective (stride {fps_stride})")
    print(f"Duration   : {duration:.1f}s  ({n_raw} frames)")
    print(f"Max clips  : ~{(n_eff_est - CLIP_LEN) // STRIDE}")
    print("Streaming frames (face crop cached every 15 frames)...\n")

    # Rolling buffer — stores 128×128 uint8, uses ~8 MB regardless of video length
    buf            : deque = deque(maxlen=CLIP_LEN + 1)
    results        : list[tuple[float, float]] = []
    last_crop_box  = None   # (x1, y1, x2, y2) in downscaled coords
    scale          = 1.0    # downsample scale applied before Haar

    raw_idx = 0
    eff_idx = 0

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        # FPS subsampling
        if raw_idx % fps_stride != 0:
            raw_idx += 1
            continue

        # Downsample to max 540p for fast Haar detection
        h, w = frame.shape[:2]
        if h > 540:
            scale  = 540.0 / h
            proc   = cv2.resize(frame, (int(w * scale), 540))
        else:
            scale  = 1.0
            proc   = frame

        # Haar face detect every 15 effective frames; cache bbox between
        if det is not None and eff_idx % 15 == 0:
            gray  = cv2.cvtColor(proc, cv2.COLOR_BGR2GRAY)
            faces = det.detectMultiScale(
                gray, scaleFactor=1.1, minNeighbors=5, minSize=(40, 40)
            )
            if len(faces) > 0:
                fx, fy, fw, fh = faces[0]
                ph, pw   = int(fh * 0.3), int(fw * 0.3)
                ph_small, pw_small = proc.shape[:2]
                last_crop_box = (
                    max(0, fx - pw), max(0, fy - ph),
                    min(pw_small, fx + fw + pw), min(ph_small, fy + fh + ph),
                )

        # Crop to face region (or full frame if no detection yet)
        if last_crop_box is not None:
            x1, y1, x2, y2 = last_crop_box
            roi = proc[y1:y2, x1:x2]
            roi = roi if roi.size > 0 else proc
        else:
            roi = proc

        # Resize to model input size — this is all we keep in memory
        buf.append(cv2.resize(roi, (INPUT_SIZE, INPUT_SIZE)))

        # Run inference when buffer is full and we're at the right stride position
        if len(buf) == CLIP_LEN + 1 and (eff_idx - CLIP_LEN) % STRIDE == 0:
            frames_f = np.array(buf, dtype=np.float32)                 # (161, 128, 128, 3)
            lum      = frames_f.mean(axis=(1, 2, 3), keepdims=True)
            frames_f /= (lum + 1e-6)
            diff = (frames_f[1:] - frames_f[:-1]) / (frames_f[1:] + frames_f[:-1] + 1e-6)
            diff = np.clip(diff, -3.0, 3.0).transpose(0, 3, 1, 2)[np.newaxis]  # (1,160,3,128,128)

            bvp      = sess.run(None, {in_name: diff})[0][0]
            hr       = bvp_to_hr(bvp, eff_fps)
            snr      = _bvp_snr(bvp, eff_fps)
            center_t = (eff_idx - CLIP_LEN / 2) / eff_fps
            results.append((center_t, hr, snr))

        eff_idx += 1
        raw_idx += 1

        if eff_idx % 150 == 0:
            pct = eff_idx / n_eff_est * 100
            n_clips = len(results)
            print(f"  {pct:5.1f}%  ({eff_idx}/{n_eff_est} frames, {n_clips} clips done)", end="\r", flush=True)

    cap.release()
    print(f"\nDone — {len(results)} clips analysed.\n")

    # ── Temporal continuity filter then sliding-window median (5 clips ≈ 13 s)
    # Reject clips where HR jumps >25 BPM from the rolling 5-clip median.
    # ref is computed from hrs_filt (filtered history) so a burst can't
    # contaminate the reference for subsequent clips.
    # WARMUP_CLIPS: pass through raw during model settle-time (~20 s).
    # By clip 8 the buffer holds [60, 80, 122, 65, 82] → median 80 → safe seed.
    JUMP_THRESH   = 25.0
    WARMUP_CLIPS  = 8
    hrs_raw  = [r[1] for r in results]
    hrs_filt = []
    for i, hr in enumerate(hrs_raw):
        if i < WARMUP_CLIPS:
            hrs_filt.append(hr)
            continue
        lo  = max(0, i - 5)
        ref = float(np.median(hrs_filt[lo:i]))   # causal: uses filtered history
        hrs_filt.append(hr if abs(hr - ref) <= JUMP_THRESH else ref)

    smoothed = []
    W = 5
    for i in range(len(hrs_filt)):
        lo = max(0, i - W // 2)
        hi = min(len(hrs_filt), lo + W)
        smoothed.append(float(np.median(hrs_filt[lo:hi])))

    # ── Report ────────────────────────────────────────────────────────────────
    GT_WINDOW  = 20.0   # match a clip to GT if within this many seconds
    SNR_THRESH = 3.5

    print(f"{'Time':>7}  {'Raw HR':>7}  {'Smooth':>7}  {'SNR':>5}  {'Oximeter':>9}  {'Err(raw)':>9}  {'Err(smo)':>9}")
    print(f"{'-'*68}")

    errors_raw, errors_snr, errors_smo = [], [], []
    csv_rows = []

    for i, (t, hr, snr) in enumerate(results):
        smo = smoothed[i]
        gt_hr = None
        if ground_truth:
            nearest = min(ground_truth, key=lambda x: abs(t - x[0]))
            if abs(t - nearest[0]) <= GT_WINDOW:
                gt_hr = nearest[1]

        mins, secs = int(t // 60), int(t % 60)
        gt_str   = f"{gt_hr:.0f}" if gt_hr is not None else "—"
        snr_str  = f"{snr:.1f}"
        err_raw  = err_smo = ""

        if gt_hr is not None:
            e_raw = hr  - gt_hr
            e_smo = smo - gt_hr
            err_raw = f"{e_raw:+.1f}"
            err_smo = f"{e_smo:+.1f}"
            errors_raw.append(abs(e_raw))
            errors_smo.append(abs(e_smo))
            if snr >= SNR_THRESH:
                errors_snr.append(abs(e_raw))

        flag = " ✓" if snr >= SNR_THRESH else "  "
        print(f"{mins:2d}m{secs:02d}s  {hr:7.1f}  {smo:7.1f}  {snr_str:>5}{flag}  {gt_str:>9}  {err_raw:>9}  {err_smo:>9}")
        gt_out  = gt_hr  if gt_hr  is not None else ""
        err_out = (hr - gt_hr) if gt_hr is not None else ""
        csv_rows.append(f"{t:.1f},{hr:.1f},{smo:.1f},{snr:.2f},{gt_out},{err_out}\n")

    print(f"{'-'*68}")
    if errors_raw:
        print(f"\nMAE  all clips        : {np.mean(errors_raw):.1f} ± {np.std(errors_raw):.1f} BPM  (n={len(errors_raw)})")
    if errors_snr:
        print(f"MAE  SNR ≥ {SNR_THRESH} only  : {np.mean(errors_snr):.1f} ± {np.std(errors_snr):.1f} BPM  (n={len(errors_snr)})  ← matches live system")
    if errors_smo:
        print(f"MAE  5-clip smoothed  : {np.mean(errors_smo):.1f} ± {np.std(errors_smo):.1f} BPM  (n={len(errors_smo)})")
    n_good = sum(1 for _, _, s in results if s >= SNR_THRESH)
    print(f"\nClips above SNR threshold: {n_good}/{len(results)} ({100*n_good/len(results):.0f}%)")

    out_csv = Path("analysis_output.csv")
    with open(out_csv, "w") as f:
        f.write("center_time_s,pred_hr,smoothed_hr,snr,oximeter_hr,error\n")
        f.writelines(csv_rows)
    print(f"Saved → {out_csv}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def _parse_gt(s: str) -> list[tuple[float, float]]:
    gt = []
    for part in s.split(","):
        t_str, bpm_str = part.strip().split(":")
        t_str = t_str.strip()
        if "m" in t_str:
            m, rest = t_str.split("m")
            secs = int(m) * 60 + int(rest.replace("s", "") or 0)
        else:
            secs = float(t_str.replace("s", ""))
        gt.append((float(secs), float(bpm_str)))
    return sorted(gt)


def _parse_gt_xlsx(path: str) -> list[tuple[float, float]]:
    """
    Read GT from an Excel file.  Expects two columns (any order, case-insensitive):
      - a time/timestamp column  (seconds as float, or MM:SS string, or 0m00s string)
      - an HR/heart rate column  (BPM as float)
    Column names are auto-detected by keyword matching.
    Requires:  pip install openpyxl
    """
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl required for --gt-xlsx.  Run:  pip install openpyxl")

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [str(c.value).lower().strip() if c.value else "" for c in ws[1]]

    # Find time column
    time_col = next(
        (i for i, h in enumerate(headers) if any(k in h for k in ("time", "timestamp", "t ", "sec"))),
        0,
    )
    # Find HR column
    hr_col = next(
        (i for i, h in enumerate(headers) if any(k in h for k in ("hr", "heart", "bpm", "rate"))),
        1,
    )
    print(f"GT Excel: time=col {time_col+1} ({headers[time_col]!r}), HR=col {hr_col+1} ({headers[hr_col]!r})")

    gt = []
    for row in list(ws.iter_rows(values_only=True))[1:]:
        t_raw, hr_raw = row[time_col], row[hr_col]
        if t_raw is None or hr_raw is None:
            continue
        t_str = str(t_raw).strip()
        # Parse various time formats: float seconds, "1:23", "1m23s", "0m30s"
        try:
            if "m" in t_str:
                m, rest = t_str.split("m")
                secs = int(m) * 60 + int(rest.replace("s", "") or 0)
            elif ":" in t_str:
                parts = t_str.split(":")
                secs = int(parts[0]) * 60 + float(parts[1])
            else:
                secs = float(t_str)
            gt.append((float(secs), float(hr_raw)))
        except (ValueError, AttributeError):
            continue
    return sorted(gt)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--video",     required=True)
    p.add_argument("--onnx-path", default=None)
    p.add_argument("--gt",        default=None,
                   help="Ground truth string: 'sec:bpm,...' or '1m18s:79,...'")
    p.add_argument("--gt-xlsx",   default=None,
                   help="Ground truth Excel file (auto-detects time + HR columns)")
    args = p.parse_args()

    onnx_path = Path(args.onnx_path) if args.onnx_path else ONNX_PATH
    if not onnx_path.exists():
        sys.exit(f"ONNX not found: {onnx_path}")

    if args.gt_xlsx:
        gt = _parse_gt_xlsx(args.gt_xlsx)
    elif args.gt:
        gt = _parse_gt(args.gt)
    else:
        gt = None
    analyze(args.video, gt, onnx_path)


if __name__ == "__main__":
    main()
